"""
Test Phase 5 — Invoice Engine (PDF + ERP Excel)
Run: venv/bin/python tests/test_invoice_engine.py

Generates real PDF and Excel files into output/pdf/ and output/excel/
"""
import sys
import json
from pathlib import Path
from datetime import date

sys.path.insert(0, str(Path(__file__).parent.parent))

from models.schema import Employee, Client, Contract, TimesheetEntry, ExtractedDocument
from models.validation import EngineResult
from engines.processing_engine import ProcessingEngine
from engines.validation_engine import ValidationEngine
from engines.invoice_engine import InvoiceEngine

print("Testing Phase 5 — Invoice Engine...")
print("=" * 55)

# ── Helper: build upstream results ────────────────────────────────────────────
def make_upstream(
    employee_id="EMP10001", employee_name="Carlos Smith",
    client_id="CL001", client_name="Emirates Steel Industries LLC",
    billing_rate=50.78, currency="AED",
    hours_per_day=None, contracted=192.0,
    gst=False,
):
    if hours_per_day is None:
        hours_per_day = [8] * 24

    emp = Employee(employee_id=employee_id, name=employee_name,
                   designation="Software Engineer")
    cli = Client(client_id=client_id, company_name=client_name,
                 country="UAE", currency=currency)
    con = Contract(
        contract_id=f"CON-{employee_id}-{client_id}",
        client_id=client_id, employee_id=employee_id,
        billing_rate=billing_rate, currency=currency,
        contracted_hours=contracted,
        start_date=date(2026, 1, 1), end_date=date(2026, 12, 31),
        overtime_allowed=True, overtime_multiplier=1.5,
        early_completion_policy="pay_actual",
        gst_applicable=gst, gst_rate=0.18 if gst else 0.0,
        payment_terms_days=30,
    )
    timesheet = [
        TimesheetEntry(date=date(2026, 6, i+1), employee_id=employee_id,
                       hours_worked=h)
        for i, h in enumerate(hours_per_day)
        if date(2026, 6, i+1).weekday() < 6
    ]
    doc = ExtractedDocument(
        employee=emp, client=cli, contract=con, timesheet=timesheet,
        billing_period_start=date(2026, 6, 1),
        billing_period_end=date(2026, 6, 30),
    )
    doc_result = EngineResult(stage="document", status="SUCCESS", confidence=0.97)
    doc_result.data = json.loads(doc.model_dump_json())
    proc_result = ProcessingEngine.process(doc_result)
    val_result  = ValidationEngine.process(doc_result, proc_result)
    return doc_result, proc_result, val_result


# ── Test 1: Standard clean invoice (AED, no GST) ─────────────────────────────
print("\nTest 1: Standard invoice — Carlos Smith (AED, no GST)")
dr, pr, vr = make_upstream()
result = InvoiceEngine.process(dr, pr, vr, sequence=1)

assert result.status == "SUCCESS", f"Expected SUCCESS: {result.errors}"
inv = result.data
assert inv["invoice_number"].startswith("INV-")
assert inv["employee_name"] == "Carlos Smith"
assert inv["client_name"] == "Emirates Steel Industries LLC"
assert inv["billing"]["currency"] == "AED"
assert inv["billing"]["gst_amount"] == 0.0
assert inv["pdf_path"] is not None
assert inv["excel_path"] is not None
assert Path(inv["pdf_path"]).exists(), "PDF file not created"
assert Path(inv["excel_path"]).exists(), "Excel file not created"

print(f"  ✓ Invoice number  : {inv['invoice_number']}")
print(f"  ✓ Total           : {inv['billing']['currency']} {inv['billing']['total_amount']:,.2f}")
print(f"  ✓ Total (INR)     : ₹{inv['billing']['total_amount_inr']:,.2f}")
print(f"  ✓ PDF             : {Path(inv['pdf_path']).name}")
print(f"  ✓ ERP Excel       : {Path(inv['excel_path']).name}")

# ── Test 2: Invoice with overtime ─────────────────────────────────────────────
print("\nTest 2: Invoice with overtime — Ahmed Khan")
hours = [10] * 24   # 240h total vs 192h contracted → 48h OT
dr2, pr2, vr2 = make_upstream(
    employee_id="EMP10002", employee_name="Ahmed Khan",
    billing_rate=47.28, contracted=192.0, hours_per_day=hours,
)
result2 = InvoiceEngine.process(dr2, pr2, vr2, sequence=2)
assert result2.status == "SUCCESS"
inv2 = result2.data
assert inv2["billing"]["overtime_hours"] > 0
print(f"  ✓ Invoice number  : {inv2['invoice_number']}")
print(f"  ✓ Overtime hours  : {inv2['billing']['overtime_hours']}h")
print(f"  ✓ OT amount       : {inv2['billing']['currency']} {inv2['billing']['overtime_amount']:,.2f}")
print(f"  ✓ Total           : {inv2['billing']['currency']} {inv2['billing']['total_amount']:,.2f}")
print(f"  ✓ PDF             : {Path(inv2['pdf_path']).name}")

# ── Test 3: Invoice with GST — build BillingResult directly ──────────────────
print("\nTest 3: Invoice with GST — INR billing")
from models.invoice import Invoice, BillingResult, LineItem
from models.validation import EngineResult as ER

billing_gst = BillingResult(
    regular_hours=160.0, overtime_hours=0.0,
    regular_amount=240000.0, overtime_amount=0.0,
    subtotal=240000.0, gst_amount=43200.0,
    total_amount=283200.0, currency="INR",
    exchange_rate_to_inr=1.0, total_amount_inr=283200.0,
    line_items=[LineItem(description="Regular hours (160h × ₹1500/hr)",
                         hours=160.0, rate=1500.0, amount=240000.0)],
    billing_notes=["GST @ 18% = 43200"],
)
inv3_obj = Invoice(
    invoice_number="INV-2026-INR001-0003",
    invoice_date=date(2026, 7, 1), due_date=date(2026, 7, 31),
    employee_id="EMP_INR", employee_name="Test Employee INR",
    client_id="INR001", client_name="Infosys Ltd",
    contract_id="CON_INR",
    billing_period_start=date(2026, 6, 1), billing_period_end=date(2026, 6, 30),
    billing=billing_gst,
)
from services.invoice_service import generate_pdf
from services.export_service import generate_erp_excel
pdf3   = generate_pdf(inv3_obj)
excel3 = generate_erp_excel(inv3_obj)
assert pdf3.exists()
assert excel3.exists()
print(f"  ✓ Invoice number  : {inv3_obj.invoice_number}")
print(f"  ✓ Subtotal        : ₹{billing_gst.subtotal:,.2f}")
print(f"  ✓ GST (18%)       : ₹{billing_gst.gst_amount:,.2f}")
print(f"  ✓ Total           : ₹{billing_gst.total_amount:,.2f}")
print(f"  ✓ PDF             : {pdf3.name}")

# ── Test 4: Validation INVALID → invoice blocked ──────────────────────────────
print("\nTest 4: Blocked when validation fails")
from models.validation import EngineResult as ER
bad_val = ER(stage="validation", status="FAILED")
bad_val.data = {"overall": "INVALID", "total_checks": 1,
                "passed": 0, "failed": 1, "warnings": 0}
bad_val.add_error("Duplicate invoice detected")
result4 = InvoiceEngine.process(dr, pr, bad_val, sequence=99)
assert result4.status == "FAILED"
assert result4.data is None
print(f"  ✓ Invoice blocked — {result4.errors[0][:55]}")

# ── Test 5: Verify ERP Excel structure ────────────────────────────────────────
print("\nTest 5: ERP Excel sheet structure")
import openpyxl
wb = openpyxl.load_workbook(inv["excel_path"])
assert "Invoice Summary" in wb.sheetnames
assert "Line Items" in wb.sheetnames
assert "ERP Upload" in wb.sheetnames
erp_ws = wb["ERP Upload"]
headers = [erp_ws.cell(1, c).value for c in range(1, 5)]
assert "COMPANY_CODE" in headers
assert "VENDOR_CODE" in headers
print(f"  ✓ Sheets          : {wb.sheetnames}")
print(f"  ✓ ERP headers     : {headers[:4]}")
print(f"  ✓ ERP rows        : {erp_ws.max_row - 1} line item(s)")

# ── Summary ────────────────────────────────────────────────────────────────────
print("\n" + "=" * 55)
print("  OUTPUT FILES")
print("=" * 55)
from config import PDF_OUTPUT_DIR, EXCEL_OUTPUT_DIR
pdfs   = list(PDF_OUTPUT_DIR.glob("*.pdf"))
excels = list(EXCEL_OUTPUT_DIR.glob("*.xlsx"))
for f in pdfs:
    print(f"  PDF   → {f.name}")
for f in excels:
    print(f"  Excel → {f.name}")

print("\nPASS — Phase 5: Invoice Engine")
