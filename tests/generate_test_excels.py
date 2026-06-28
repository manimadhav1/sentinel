"""
Generate 4 test Excel sheets with varying confidence levels.
Run: venv/bin/python tests/generate_test_excels.py
Output: data/sample_docs/
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("data/sample_docs")
OUT.mkdir(parents=True, exist_ok=True)

# ── Style helpers ──────────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=11):
    return Font(bold=True, size=size)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(wb, title, headers, rows, header_color="1B2A4A", notes=None):
    ws = wb.active if wb.worksheets else wb.create_sheet()
    ws.title = title

    # Notes row at top (context Gemini will read)
    row_offset = 0
    if notes:
        for i, note in enumerate(notes):
            ws.cell(row=i+1, column=1, value=note).font = Font(italic=True, color="666666", size=9)
        row_offset = len(notes) + 1

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_offset + 1, column=col, value=h)
        cell.font    = Font(bold=True, size=10, color="FFFFFF")
        cell.fill    = hdr_fill(header_color)
        cell.alignment = Alignment(horizontal="center")
        cell.border  = thin_border()

    # Data rows
    for r, row in enumerate(rows, row_offset + 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    # Auto width
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 40)

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ~90% confidence
# Clear standard payroll with all key fields present and labelled properly.
# Gemini will find employee_id, customer_code, days, salary, period easily.
# ══════════════════════════════════════════════════════════════════════════════
wb1 = openpyxl.Workbook()
write_sheet(
    wb1, "Payroll_June2026",
    headers=["Employee_ID", "Employee_Name", "Customer_Code",
             "Working_Days", "Basic_Salary_AED", "OT_Hours", "Pay_Period"],
    rows=[
        ["EMP003", "James Okonkwo",    "CUST002", 22, 11000, 4,  "June 2026"],
        ["EMP005", "Omar Al Hashimi",  "CUST003", 26, 14000, 0,  "June 2026"],
        ["EMP008", "Aisha Bint Khalid","CUST004", 20, 12500, 6,  "June 2026"],
    ],
    notes=[
        "TASC Outsourcing — Monthly Payroll Summary",
        "Period: June 2026 (01-Jun-2026 to 30-Jun-2026)",
        "Currency: AED | Billing Type: Hourly",
    ],
    header_color="1B4F72",
)
p1 = OUT / "test_90pct_clear_payroll.xlsx"
wb1.save(str(p1))
print(f"✓ ~90% confidence: {p1.name}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ~85% confidence
# Non-standard column names and no explicit pay period — Gemini has to infer.
# Employee IDs present but customer reference is a name not a code.
# ══════════════════════════════════════════════════════════════════════════════
wb2 = openpyxl.Workbook()
write_sheet(
    wb2, "Staff_Hours_Report",
    headers=["Staff_No", "Full_Name", "Account_Code",
             "Days_Attended", "Monthly_CTC", "OT_Days"],
    rows=[
        ["EMP010", "Meera Pillai",      "CUST005", 21, 9200,  2],
        ["EMP011", "Hassan Al Muhairi", "CUST006", 25, 10500, 5],
        ["EMP013", "Vikram Singh",      "CUST007", 23, 13000, 0],
    ],
    notes=[
        "HR System Export — June Monthly Closing",
        "All amounts in local currency (AED)",
        "Column names differ from standard payroll format — AI must infer mapping",
    ],
    header_color="1A5276",
)
p2 = OUT / "test_85pct_nonstandard_columns.xlsx"
wb2.save(str(p2))
print(f"✓ ~85% confidence: {p2.name}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — ~60% confidence → human review
# No employee IDs (just names), no customer codes, no explicit billing period,
# compensation given as a weekly figure, ambiguous client names.
# ══════════════════════════════════════════════════════════════════════════════
wb3 = openpyxl.Workbook()
write_sheet(
    wb3, "Weekly_Hours",
    headers=["Worker_Name", "Project_Site", "Week1_Days", "Week2_Days",
             "Week3_Days", "Week4_Days", "Weekly_Pay"],
    rows=[
        ["Rahul M.",        "Steel plant - AD",  5, 5, 4, 5, 2125],
        ["Nadia Z.",        "Contracting Co.",   4, 5, 5, 4, 1800],
        ["Unknown Worker",  "Site B",            3, 4, 3, 4, 1500],
    ],
    notes=[
        "Weekly attendance report — Q2 summary",
        "No employee codes — refer to HR system for IDs",
        "Payment weekly. Client site names used, not codes.",
    ],
    header_color="7D6608",
)
p3 = OUT / "test_60pct_human_review.xlsx"
wb3.save(str(p3))
print(f"✓ ~60% confidence (review): {p3.name}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ~45% confidence → high priority human review
# Completely non-standard: mixed currencies, no IDs anywhere, dates vague,
# compensation is total cost-to-company (not hourly/daily rate).
# Multiple employees on one row (group billing). Very hard for AI to parse.
# ══════════════════════════════════════════════════════════════════════════════
wb4 = openpyxl.Workbook()
write_sheet(
    wb4, "Cost_Allocation",
    headers=["Resource_Group", "Engagement", "Headcount",
             "Total_Days_Pool", "CTC_USD", "Notes"],
    rows=[
        ["Offshore Dev Team", "Project Alpha",  3, 55, 18500, "Mix of senior/junior. AED equivalent TBD."],
        ["QA Resources",      "Project Beta",   2, 40, 9200,  "Contract renewal pending. Period unclear."],
        ["Support Staff",     "Ops - General",  4, 72, 14000, "Days spread across May-June. Split unclear."],
    ],
    notes=[
        "Cost Allocation Sheet — Finance Dept",
        "USD amounts. Conversion rate not specified.",
        "Headcount billing — individual IDs not listed.",
        "Period: approx Q2 2026 (exact dates unknown)",
    ],
    header_color="6E2C00",
)
p4 = OUT / "test_45pct_high_priority_review.xlsx"
wb4.save(str(p4))
print(f"✓ ~45% confidence (high priority review): {p4.name}")

print(f"\nAll 4 files saved to: {OUT.resolve()}")
print("""
Expected results when uploaded:
  test_90pct_clear_payroll.xlsx        → Auto invoice generated
  test_85pct_nonstandard_columns.xlsx  → Auto invoice generated (with warnings)
  test_60pct_human_review.xlsx         → Routed to review queue
  test_45pct_high_priority_review.xlsx → High priority review queue
""")
