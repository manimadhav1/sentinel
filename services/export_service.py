from __future__ import annotations
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from models.invoice import Invoice
from utils.logger import get_logger
from config import EXCEL_OUTPUT_DIR

logger = get_logger("export_service")

# ── Colours ────────────────────────────────────────────────────────────────────
NAVY   = "1B2A4A"
BLUE   = "2563EB"
LIGHT  = "F1F5F9"
WHITE  = "FFFFFF"
GREY   = "CBD5E1"
GREEN  = "16A34A"


def generate_erp_excel(invoice: Invoice) -> Path:
    """
    Generate SAP-compatible ERP Excel export.
    Returns path to the saved .xlsx file.
    """
    filename = f"{invoice.invoice_number}_ERP.xlsx"
    output_path = EXCEL_OUTPUT_DIR / filename

    wb = openpyxl.Workbook()

    _build_invoice_sheet(wb, invoice)
    _build_line_items_sheet(wb, invoice)
    _build_erp_upload_sheet(wb, invoice)

    # remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(output_path))
    logger.info(f"ERP Excel generated: {output_path}")
    return output_path


# ── Sheet 1: Invoice Summary ───────────────────────────────────────────────────

def _build_invoice_sheet(wb: openpyxl.Workbook, invoice: Invoice) -> None:
    ws = wb.create_sheet("Invoice Summary")
    b  = invoice.billing

    _set_col_widths(ws, [28, 42])

    # Title banner
    ws.merge_cells("A1:B1")
    _write(ws, "A1", "SENTINEL — Invoice Summary",
           bold=True, size=14, fg=WHITE, bg=NAVY, align="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    _write(ws, "A2", "Touchless Invoice Automation",
           size=9, fg=BLUE, align="center")

    _blank(ws, 3)

    # Header fields
    fields = [
        ("Invoice Number",    invoice.invoice_number),
        ("Invoice Date",      str(invoice.invoice_date)),
        ("Due Date",          str(invoice.due_date)),
        ("Status",            invoice.status),
        ("",                  ""),
        ("Employee ID",       invoice.employee_id),
        ("Employee Name",     invoice.employee_name),
        ("Client ID",         invoice.client_id),
        ("Client Name",       invoice.client_name),
        ("Contract Ref",      invoice.contract_id),
        ("",                  ""),
        ("Billing Period",    f"{invoice.billing_period_start} → {invoice.billing_period_end}"),
        ("",                  ""),
        ("Regular Hours",     f"{b.regular_hours}h"),
        ("Overtime Hours",    f"{b.overtime_hours}h"),
        ("Regular Amount",    f"{b.currency} {b.regular_amount:,.2f}"),
        ("Overtime Amount",   f"{b.currency} {b.overtime_amount:,.2f}"),
        ("Subtotal",          f"{b.currency} {b.subtotal:,.2f}"),
        ("GST",               f"{b.currency} {b.gst_amount:,.2f}"),
        ("TOTAL DUE",         f"{b.currency} {b.total_amount:,.2f}"),
        ("Total (INR)",       f"₹{b.total_amount_inr:,.2f}"),
        ("Exchange Rate",     f"1 {b.currency} = ₹{b.exchange_rate_to_inr}"),
    ]

    for i, (label, value) in enumerate(fields, start=4):
        if not label:
            ws.row_dimensions[i].height = 8
            continue
        _write(ws, f"A{i}", label, bold=True,
               bg=LIGHT if i % 2 == 0 else WHITE, fg=NAVY)
        _write(ws, f"B{i}", value,
               bg=LIGHT if i % 2 == 0 else WHITE,
               bold=(label == "TOTAL DUE"),
               fg=BLUE if label == "TOTAL DUE" else NAVY)

    # Billing notes
    row = len(fields) + 5
    _write(ws, f"A{row}", "Billing Notes", bold=True, fg=WHITE, bg=NAVY)
    ws.merge_cells(f"A{row}:B{row}")
    for note in b.billing_notes:
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        _write(ws, f"A{row}", f"• {note}", size=9, fg=NAVY)

    _apply_outer_border(ws, f"A1:B{row}")


# ── Sheet 2: Line Items ────────────────────────────────────────────────────────

def _build_line_items_sheet(wb: openpyxl.Workbook, invoice: Invoice) -> None:
    ws = wb.create_sheet("Line Items")
    b  = invoice.billing

    _set_col_widths(ws, [50, 12, 18, 18])

    ws.merge_cells("A1:D1")
    _write(ws, "A1", f"Line Items — {invoice.invoice_number}",
           bold=True, size=12, fg=WHITE, bg=NAVY, align="center")
    ws.row_dimensions[1].height = 24

    headers = ["Description", "Hours", "Rate", "Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for row_i, item in enumerate(b.line_items, start=3):
        bg = LIGHT if row_i % 2 == 0 else WHITE
        _write(ws, f"A{row_i}", item.description, bg=bg, fg=NAVY)
        _write(ws, f"B{row_i}", item.hours or 0, bg=bg, fg=NAVY, align="right")
        _write(ws, f"C{row_i}", item.rate, bg=bg, fg=NAVY, align="right",
               num_fmt=f'"{b.currency} "#,##0.00')
        _write(ws, f"D{row_i}", item.amount, bg=bg, fg=NAVY, align="right",
               num_fmt=f'"{b.currency} "#,##0.00',
               bold=(item.amount < 0))

    total_row = len(b.line_items) + 3
    ws.merge_cells(f"A{total_row}:C{total_row}")
    _write(ws, f"A{total_row}", "TOTAL", bold=True, fg=WHITE, bg=NAVY, align="right")
    _write(ws, f"D{total_row}", b.total_amount, bold=True,
           fg=WHITE, bg=NAVY, align="right",
           num_fmt=f'"{b.currency} "#,##0.00')

    _apply_outer_border(ws, f"A1:D{total_row}")


# ── Sheet 3: ERP Upload (SAP-style) ───────────────────────────────────────────

def _build_erp_upload_sheet(wb: openpyxl.Workbook, invoice: Invoice) -> None:
    """
    SAP-compatible flat upload format.
    Each row = one line item with full invoice context.
    """
    ws = wb.create_sheet("ERP Upload")
    b  = invoice.billing

    headers = [
        "COMPANY_CODE", "VENDOR_CODE", "DOCUMENT_DATE", "POSTING_DATE",
        "DOCUMENT_TYPE", "REFERENCE", "HEADER_TEXT",
        "GL_ACCOUNT", "COST_CENTER", "PROFIT_CENTER",
        "CURRENCY", "AMOUNT", "TAX_CODE", "TAX_AMOUNT",
        "ASSIGNMENT", "ITEM_TEXT",
        "EMPLOYEE_ID", "CLIENT_ID", "CONTRACT_ID",
        "BILLING_PERIOD_START", "BILLING_PERIOD_END",
        "HOURS", "RATE", "INVOICE_NUMBER",
    ]

    _set_col_widths(ws, [14] * len(headers))

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=9)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30

    # One row per line item
    for row_i, item in enumerate(b.line_items, start=2):
        tax_code   = "V0" if b.gst_amount == 0 else "V5"
        tax_amount = round(item.amount * (b.gst_amount / b.subtotal), 2) \
                     if b.subtotal and b.gst_amount else 0.0
        bg = LIGHT if row_i % 2 == 0 else WHITE

        row_data = [
            "1000",                              # COMPANY_CODE
            invoice.client_id,                   # VENDOR_CODE
            str(invoice.invoice_date),           # DOCUMENT_DATE
            str(invoice.invoice_date),           # POSTING_DATE
            "RE",                                # DOCUMENT_TYPE (vendor invoice)
            invoice.invoice_number,              # REFERENCE
            f"Invoice {invoice.invoice_number}", # HEADER_TEXT
            "500000",                            # GL_ACCOUNT (services received)
            f"CC-{invoice.client_id}",           # COST_CENTER
            f"PC-{invoice.employee_id}",         # PROFIT_CENTER
            b.currency,                          # CURRENCY
            item.amount,                         # AMOUNT
            tax_code,                            # TAX_CODE
            tax_amount,                          # TAX_AMOUNT
            invoice.invoice_number,              # ASSIGNMENT
            item.description,                    # ITEM_TEXT
            invoice.employee_id,                 # EMPLOYEE_ID
            invoice.client_id,                   # CLIENT_ID
            invoice.contract_id,                 # CONTRACT_ID
            str(invoice.billing_period_start),   # BILLING_PERIOD_START
            str(invoice.billing_period_end),     # BILLING_PERIOD_END
            item.hours or 0,                     # HOURS
            item.rate,                           # RATE
            invoice.invoice_number,              # INVOICE_NUMBER
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font      = Font(size=9, color=NAVY)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _write(ws, cell_ref: str, value, bold=False, size=10,
           fg=NAVY, bg=WHITE, align="left", num_fmt=None) -> None:
    cell            = ws[cell_ref]
    cell.value      = value
    cell.font       = Font(bold=bold, size=size, color=fg)
    cell.fill       = PatternFill("solid", fgColor=bg)
    cell.alignment  = Alignment(horizontal=align, vertical="center",
                                wrap_text=False)
    if num_fmt:
        cell.number_format = num_fmt


def _blank(ws, row: int) -> None:
    ws.row_dimensions[row].height = 6


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _apply_outer_border(ws, cell_range: str) -> None:
    thin = Side(style="thin", color=GREY)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )
